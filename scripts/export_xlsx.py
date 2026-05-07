
"""将 v3 全部 JSON 数据导出为 xlsx，每个实体类型一个 sheet。"""

import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
OUTPUT = os.path.join(os.path.dirname(__file__), '..', 'output', 'v3_data.xlsx')

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
ZEBRA_1 = PatternFill('solid', fgColor='FFFFFF')
ZEBRA_2 = PatternFill('solid', fgColor='F7F9FC')
THIN_BORDER = Border(
    left=Side('thin', color='D9DEE7'), right=Side('thin', color='D9DEE7'),
    top=Side('thin', color='D9DEE7'), bottom=Side('thin', color='D9DEE7'),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')

def flatten(obj, prefix=''):
    if isinstance(obj, dict):
        out = {}
        for k, v in obj.items():
            key = f'{prefix}.{k}' if prefix else k
            if isinstance(v, dict):
                out.update(flatten(v, key))
            elif isinstance(v, list):
                out[key] = '|'.join(
                    json.dumps(x, ensure_ascii=False) if isinstance(x, (dict, list)) else str(x)
                    for x in v
                ) if v else ''
            else:
                out[key] = v
        return out
    return obj

def write_sheet(wb, name, rows, col_order=None):
    if not rows:
        return
    all_keys = set()
    for r in rows:
        all_keys.update(r.keys())
    if col_order:
        ordered = [k for k in col_order if k in all_keys]
        ordered += sorted(k for k in all_keys if k not in col_order)
    else:
        ordered = sorted(all_keys)

    ws = wb.create_sheet(title=name)
    for c, h in enumerate(ordered, 1):
        cell = ws.cell(1, c, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    for ri, row in enumerate(rows, 2):
        fill = ZEBRA_1 if (ri % 2 == 0) else ZEBRA_2
        for ci, key in enumerate(ordered, 1):
            val = row.get(key, '')
            if isinstance(val, dict):
                val = json.dumps(val, ensure_ascii=False)
            elif isinstance(val, list):
                val = '|'.join(str(x) for x in val) if val else ''
            cell = ws.cell(ri, ci, val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
    for ci, key in enumerate(ordered, 1):
        max_len = len(str(key))
        for ri in range(2, min(len(rows) + 2, 20)):
            v = ws.cell(ri, ci).value
            if v:
                max_len = max(max_len, min(len(str(v)), 40))
        ws.column_dimensions[get_column_letter(ci)].width = max_len + 4
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(ordered))}{len(rows) + 1}'

def main():
    wb = Workbook()
    wb.remove(wb.active)

    # 元数据
    with open(os.path.join(DATA_DIR, 'meta.json'), 'r', encoding='utf-8') as f:
        meta = json.load(f)
    write_sheet(wb, '元数据', [flatten(meta)])

    # 标签
    with open(os.path.join(DATA_DIR, 'tags.json'), 'r', encoding='utf-8') as f:
        tags_data = json.load(f)
    write_sheet(wb, '标签分类', [flatten(c) for c in tags_data.get('categories', [])],
                col_order=['id', 'name', 'description'])
    write_sheet(wb, '标签', [flatten(t) for t in tags_data.get('tags', [])],
                col_order=['id', 'name', 'category', 'description'])

    # 实体
    sheets = [
        ('项目', 'attractions.json', 'attractions',
         ['id', 'name', 'ranking', 'rating', 'zone_ids', 'tags', 'subtype',
          'duration_minutes', 'review_ids', 'opinion_ids', 'warning_ids']),
        ('演出', 'shows.json', 'shows',
         ['id', 'name', 'zone_ids', 'tags', 'location',
          'review_ids', 'opinion_ids', 'warning_ids']),
        ('餐厅', 'restaurants.json', 'restaurants',
         ['id', 'name', 'zone_ids', 'tags', 'location',
          'recommended_dish_ids', 'review_ids', 'opinion_ids', 'warning_ids']),
        ('菜品', 'dishes.json', 'dishes',
         ['id', 'name', 'restaurant_ids', 'zone_ids', 'tags',
          'review_ids', 'opinion_ids']),
        ('技巧', 'tips.json', 'tips',
         ['id', 'category', 'title', 'tags', 'attraction_ids', 'opinion_ids']),
        ('避雷', 'warnings.json', 'warnings',
         ['id', 'category', 'severity', 'content', 'alternative', 'tags',
          'zone_ids', 'attraction_ids', 'show_ids', 'restaurant_ids', 'shortcut_ids']),
        ('行程', 'itineraries.json', 'itineraries',
         ['id', 'name', 'description', 'tags']),
        ('小道', 'shortcuts.json', 'shortcuts',
         ['id', 'name', 'from', 'to', 'route', 'savings', 'tags']),
        ('评价', 'reviews.json', 'reviews',
         ['id', 'target_type', 'target_id', 'content', 'sentiment']),
        ('观点', 'opinions.json', 'opinions',
         ['id', 'target_type', 'target_id', 'stance', 'claim', 'reasoning']),
    ]
    for name, fname, key, col_order in sheets:
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fpath):
            continue
        with open(fpath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        write_sheet(wb, name, [flatten(e) for e in data.get(key, [])], col_order=col_order)

    # 行前准备
    with open(os.path.join(DATA_DIR, 'preparations.json'), 'r', encoding='utf-8') as f:
        prep = json.load(f)
    write_sheet(wb, '行前准备', [flatten(prep)])

    os.makedirs(os.path.dirname(OUTPUT), exist_ok=True)
    wb.save(OUTPUT)
    size_kb = os.path.getsize(OUTPUT) / 1024
    print(f'生成完成: {OUTPUT} ({size_kb:.1f} KB, {len(wb.sheetnames)} 个 sheet)')
    for s in wb.sheetnames:
        ws = wb[s]
        print(f'  {s}: {ws.max_row - 1} 行 × {ws.max_column} 列')

if __name__ == '__main__':
    main()
